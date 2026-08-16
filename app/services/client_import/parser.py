"""Parse `Client List for Verax.xlsx` into normalized rows (infra PRD §3.2).

Verified structure (2026-07-06): two sheets, identical 7 columns.
  - 'Client List'            (810 rows) -> writers, kind=client
  - 'Commission Partner List' (78 rows) -> writers, kind=commission_partner

Columns:
  Artist / Publisher Name | Contact Email | Contact Name | Payee Name |
  Admin Type (YT Only / MLC Only / both) | Preferred Language (EN/ES) |
  Quarterly Client?

Realities this parser absorbs, verified against the delivered file:
  - No beneficiary IDs anywhere -> matching is name-based (matcher.py).
  - Contact Email holds >=1 comma-separated addresses; the same address
    recurs across rows (managers/attorneys) -> contacts are shared.
  - Admin Type is a set spelled inconsistently ("MLC, YT ", "YT", "MLC"),
    with one anomalous "ST, YT".
  - Trailing whitespace is everywhere; every field is stripped.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from typing import List, Optional

# openpyxl is already a dependency (xlsx_parser.py uses it).
import openpyxl

CLIENT_SHEET = "Client List"
PARTNER_SHEET = "Commission Partner List"

# The 7 expected headers, normalized (lowercased, collapsed whitespace). The
# first column header differs between sheets ('Artist / Publisher Name' vs
# 'Name'); both map to the same field.
_NAME_HEADERS = {"artist / publisher name", "name"}
_EMAIL_STRICT_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Catalog tokens as written in 'Admin Type', normalized to the model's values.
_CATALOG_TOKENS = {"mlc": "MECH", "yt": "YT"}


class WriterKind(str, Enum):
    CLIENT = "client"
    COMMISSION_PARTNER = "commission_partner"


@dataclass(frozen=True)
class ParsedEmail:
    address: str
    is_valid: bool


@dataclass
class ClientRow:
    """One spreadsheet row, normalized. `row_no` is 1-based within its sheet
    (excluding the header) so findings can point back at the source."""

    sheet: str
    row_no: int
    kind: WriterKind
    name: str
    payee_name: Optional[str]
    emails: List[ParsedEmail] = field(default_factory=list)
    contact_names: List[str] = field(default_factory=list)
    # Normalized catalog codes, e.g. {"MECH", "YT"}; unknown tokens land in
    # `unknown_catalog_tokens` so the validator can flag them (C-BAD-CATALOG).
    catalogs: List[str] = field(default_factory=list)
    unknown_catalog_tokens: List[str] = field(default_factory=list)
    preferred_language: Optional[str] = None  # "en" | "es" | None
    is_quarterly: bool = False
    raw_admin_type: Optional[str] = None
    raw_quarterly: Optional[str] = None

    @property
    def valid_emails(self) -> List[str]:
        return [e.address for e in self.emails if e.is_valid]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_emails(raw: str) -> List[ParsedEmail]:
    parts = [p.strip() for p in re.split(r"[,;]", raw) if p.strip()]
    return [ParsedEmail(address=p, is_valid=bool(_EMAIL_STRICT_RE.match(p))) for p in parts]


def _split_names(raw: str) -> List[str]:
    return [p.strip() for p in re.split(r"[,;]", raw) if p.strip()]


def _parse_catalogs(raw: str):
    known: List[str] = []
    unknown: List[str] = []
    for tok in re.split(r"[,/]", raw):
        t = tok.strip().lower()
        if not t:
            continue
        if t in _CATALOG_TOKENS:
            code = _CATALOG_TOKENS[t]
            if code not in known:
                known.append(code)
        else:
            unknown.append(tok.strip())
    return known, unknown


def _parse_language(raw: str) -> Optional[str]:
    t = raw.strip().lower()
    if t.startswith("es"):
        return "es"
    if t.startswith("en"):
        return "en"
    return None


def _header_index(header_row) -> dict:
    """Map normalized header text -> column index for one sheet."""
    idx = {}
    for i, cell in enumerate(header_row):
        idx[_clean(cell).lower()] = i
    return idx


def _col(idx: dict, *candidates: str) -> Optional[int]:
    for c in candidates:
        if c in idx:
            return idx[c]
    return None


def _parse_sheet(ws, kind: WriterKind) -> List[ClientRow]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    idx = _header_index(rows[0])

    name_col = next((idx[h] for h in _NAME_HEADERS if h in idx), 0)
    email_col = _col(idx, "contact email")
    contact_name_col = _col(idx, "contact name")
    payee_col = _col(idx, "payee name")
    admin_col = _col(idx, "admin type (yt only / mlc only / both)", "admin type")
    lang_col = _col(idx, "preferred language (en/es)", "preferred language")
    quarterly_col = _col(idx, "quarterly client?", "quarterly client")

    def get(row, col):
        return _clean(row[col]) if col is not None and col < len(row) else ""

    out: List[ClientRow] = []
    for n, row in enumerate(rows[1:], start=1):
        name = get(row, name_col)
        # Skip fully-blank rows (trailing spreadsheet padding).
        if not any(_clean(v) for v in row):
            continue
        catalogs, unknown = _parse_catalogs(get(row, admin_col))
        raw_q = get(row, quarterly_col)
        out.append(
            ClientRow(
                sheet=ws.title,
                row_no=n,
                kind=kind,
                name=name,
                payee_name=get(row, payee_col) or None,
                emails=_split_emails(get(row, email_col)),
                contact_names=_split_names(get(row, contact_name_col)),
                catalogs=catalogs,
                unknown_catalog_tokens=unknown,
                preferred_language=_parse_language(get(row, lang_col)),
                is_quarterly=raw_q.lower().startswith(("yes", "both")),
                raw_admin_type=get(row, admin_col) or None,
                raw_quarterly=raw_q or None,
            )
        )
    return out


def parse_client_list(path: str) -> List[ClientRow]:
    """Parse both known sheets. Sheets are matched by title; a missing sheet
    is simply skipped (the caller surfaces that as a finding if needed)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: List[ClientRow] = []
    if CLIENT_SHEET in wb.sheetnames:
        rows.extend(_parse_sheet(wb[CLIENT_SHEET], WriterKind.CLIENT))
    if PARTNER_SHEET in wb.sheetnames:
        rows.extend(_parse_sheet(wb[PARTNER_SHEET], WriterKind.COMMISSION_PARTNER))
    wb.close()
    return rows
