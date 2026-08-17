"""XLSX detail parser (PRD §2.4).

Parses a statement detail spreadsheet (sheet ``Blad1``) into statement_line
mappings, handling the quirks verified across 2,612 real files:

- header row may not be row 1 (scan for the row whose first cell is 'Period')
- the last data row is a grand-total row (all cells None except Earnings) —
  excluded from lines, captured as ``embedded_total``; its Earnings cell is a
  formula, so the workbook is opened ``data_only=True`` to read the cached value
- two schema variants (Mechanical vs YouTube) map into the statement_line
  superset; the literal header typo 'WrtierSplit%' is real — map it, don't fix it
- identity fields (SongCode, SongTitle, ...) can be numeric in the file —
  coerced to str; money stays Decimal

Statement files are immutable records — this module only ever reads them.
"""

from decimal import Decimal
from typing import Dict, List, Optional, Tuple

import openpyxl

from app.models.statements import StatementLine

SHEET_NAME = "Blad1"

# XLSX header -> statement_line column (None = redundant per-statement metadata)
HEADER_MAP = {
    "Period": None,
    "Beneficiary": None,
    "Name": None,
    "SongCode": "song_code",  # Mechanical
    "AssetID": "asset_id",  # YouTube
    "CustomID Client": "custom_id",  # YouTube
    "SongTitle": "song_title",
    "Country": "country",
    "Channel": "channel",
    "IncomeSource": "income_source",
    "IncomeType": "income_type",
    "Price": "price",
    "CommissionRate%": "commission_pct",
    "RBP": "rbp",
    "Rate_Applied": "rate_applied",
    "WrtierSplit%": "writer_split_pct",  # Mechanical — typo is in the real header
    "ContPer": "writer_split_pct",  # YouTube — same position/meaning
    "BenSplit%": "ben_split_pct",
    "Units": "units",
    "Earnings": "earnings",
}

DECIMAL_FIELDS = {
    "price",
    "commission_pct",
    "rbp",
    "rate_applied",
    "writer_split_pct",
    "ben_split_pct",
    "units",
    "earnings",
}


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _to_decimal(value) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def parse_statement_xlsx(path) -> Tuple[List[Dict], Decimal, Optional[Decimal], int]:
    """Parse a detail XLSX into (lines, detail_sum, embedded_total, line_count).

    ``lines`` are dicts keyed by statement_line columns (with ``row_no``, without
    ``statement_id``), ready for :func:`persist_lines`.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

        columns = None  # list of statement_line field names per cell index
        earnings_idx = None
        lines: List[Dict] = []
        detail_sum = Decimal("0")
        embedded_total: Optional[Decimal] = None

        for row in ws.iter_rows(values_only=True):
            if columns is None:
                first = row[0] if row else None
                if isinstance(first, str) and first.strip() == "Period":
                    columns = [
                        HEADER_MAP.get(str(cell).strip()) if cell is not None else None
                        for cell in row
                    ]
                    earnings_idx = columns.index("earnings")
                continue

            if all(cell is None for cell in row):
                continue

            # Grand-total row: every cell None except Earnings (PRD §2.4 #1)
            if len(row) > earnings_idx and row[earnings_idx] is not None and all(
                cell is None for i, cell in enumerate(row) if i != earnings_idx
            ):
                embedded_total = _to_decimal(row[earnings_idx])
                continue

            line: Dict = {"row_no": len(lines) + 1}
            for idx, field in enumerate(columns):
                if field is None or idx >= len(row):
                    continue
                value = row[idx]
                line[field] = (
                    _to_decimal(value) if field in DECIMAL_FIELDS else _to_str(value)
                )
            lines.append(line)
            if line.get("earnings") is not None:
                detail_sum += line["earnings"]

        if columns is None:
            raise ValueError("No header row starting with 'Period' found in %s" % path)

        return lines, detail_sum, embedded_total, len(lines)
    finally:
        wb.close()


# --- fast line persistence ----------------------------------------------------
#
# statement_line is the volume table: a full half-year drop inserts ~6.7 million
# rows, and the single biggest statement carries ~1M lines on its own. Row-wise
# INSERTs (bulk_insert_mappings -> executemany) are what made ingest take hours
# on Postgres; COPY streams the same rows in one protocol exchange and is
# typically 5-10x faster on exactly the statements that dominate the run.
#
# COPY text format is used (not csv): tab-delimited, backslash-escaped, with \N
# for NULL. It is the only format where NULL vs empty-string is unambiguous —
# and these are royalty figures, so "no value" and "zero-length text" must never
# be conflated by the transport.

def _copy_encode(value) -> str:
    """One value -> COPY text-format field."""
    if value is None:
        return "\\N"
    text = str(value)
    if "\\" in text:
        text = text.replace("\\", "\\\\")
    if "\t" in text:
        text = text.replace("\t", "\\t")
    if "\n" in text:
        text = text.replace("\n", "\\n")
    if "\r" in text:
        text = text.replace("\r", "\\r")
    return text


class _RowStream:
    """File-like reader over encoded rows, so COPY streams instead of holding
    a ~200 MB payload for the biggest statement in memory."""

    def __init__(self, chunks):
        self._chunks = chunks
        self._buffer = b""

    def read(self, size=-1) -> bytes:
        while size < 0 or len(self._buffer) < size:
            try:
                self._buffer += next(self._chunks)
            except StopIteration:
                break
        if size < 0:
            out, self._buffer = self._buffer, b""
            return out
        out, self._buffer = self._buffer[:size], self._buffer[size:]
        return out

    # psycopg2 probes readline on some paths; read() is what it uses for COPY.
    def readline(self):  # pragma: no cover
        return self.read(8192)


def persist_lines(statement_id: int, lines: List[Dict], session) -> int:
    """Insert parsed lines for a statement. Flushes/joins the session's
    transaction, does not commit — the worker owns commit cadence."""
    if session.get_bind().dialect.name != "postgresql":
        # SQLite (dev, tests): COPY does not exist; the original path stays.
        session.bulk_insert_mappings(
            StatementLine,
            [dict(line, statement_id=statement_id) for line in lines],
        )
        session.flush()
        return len(lines)

    cols = [c.name for c in StatementLine.__table__.columns if c.name != "id"]

    def chunks():
        buf, size = [], 0
        for line in lines:
            row = dict(line, statement_id=statement_id)
            buf.append("\t".join(_copy_encode(row.get(c)) for c in cols) + "\n")
            size += len(buf[-1])
            if size >= 1_000_000:  # ~1 MB per network write
                yield "".join(buf).encode("utf-8")
                buf, size = [], 0
        if buf:
            yield "".join(buf).encode("utf-8")

    # Same transaction as the ORM session: the worker's commit/rollback applies
    # to these rows exactly as it did to the executemany path.
    session.flush()
    dbapi = session.connection().connection
    with dbapi.cursor() as cur:
        cur.copy_expert(
            'COPY statement_line ({}) FROM STDIN'.format(", ".join(cols)),
            _RowStream(chunks()),
        )
    return len(lines)
